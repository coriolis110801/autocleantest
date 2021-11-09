from openpyxl import load_workbook
import sqlite3

conn = sqlite3.connect("db.sqlite3")

workbook = load_workbook("1.xlsx")  # 找到需要xlsx文件的位置
booksheet = workbook.active  # 获取当前活跃的sheet,默认是第一个sheet

# 如果想获取别的sheet页采取下面这种方式，先获取所有sheet页名，在通过指定那一页。
# sheets = workbook.get_sheet_names()  # 从名称获取sheet
# booksheet = workbook.get_sheet_by_name(sheets[0])
# 获取sheet页的行数据
rows = booksheet.rows
# 获取sheet页的列数据
columns = booksheet.columns
i = 0
# 迭代所有的行
for row in rows:
    i = i + 1
    # line = [col.value for col in row]
    cell1 = booksheet.cell(row=i, column=1).value  # 获取第i行1 列的数据
    cell2 = booksheet.cell(row=i, column=2).value  # 获取第i行 2 列的数据

    cell3 = booksheet.cell(row=i, column=3).value  # 获取第i行 3 列的数据

    cell4 = booksheet.cell(row=i, column=4).value  # 获取第i行 4 列的数据
    cell5 = booksheet.cell(row=i, column=5).value  # 获取第i行 4 列的数据

    cell6 = booksheet.cell(row=i, column=6).value  # 获取第i行 4 列的数据

    #print(cell1+" "+cell2+" "+cell3+" "+cell4+" "+cell5+" "+cell6)
    sql = "insert into store_product(name,price,digital,image,category_id,pos) values('%s',%f,%d,'%s',%d,%d)" % (cell1, cell2, cell3,cell4,cell5,cell6)
    print(sql)
    conn.execute(sql)
    conn.commit()

conn.close()
print('ok');